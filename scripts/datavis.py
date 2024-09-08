import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
import os

file = 'logs/RegistrosEMEQ.xlsx'

measured_data = {
    "Data": [],
    "Hora": [],
    "Temperatura externa (°C)": [],
    "Umidade (%)": [],
    "Nível de CO (ppm)": [],
    "Temperatura da água (°C)": [],
    "pH": [],
    "Turbidez (%)": [],
    "Partículas dissolvidas (ppm)": [],
}

# Sample data to append
measured_data["Data"].append(datetime.now().strftime("%d-%m-%Y"))
measured_data["Hora"].append(datetime.now().strftime("%H:%M:%S"))
measured_data["Temperatura externa (°C)"].append(25)
measured_data["Umidade (%)"].append(60)
measured_data["Nível de CO (ppm)"].append(0.5)
measured_data["Temperatura da água (°C)"].append(20)
measured_data["pH"].append(7.5)
measured_data["Turbidez (%)"].append(0.5)
measured_data["Partículas dissolvidas (ppm)"].append(0.5)

# Convert measured_data to DataFrame

def excelLog(file, data):
    df_new = pd.DataFrame(data)

    book = load_workbook(file)

    # Read the existing data from the specific sheet
    sheet_name = 'logs'
    if sheet_name in book.sheetnames:
        df_existing = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
    else:
        df_existing = pd.DataFrame()

    # Append new data to the existing DataFrame
    df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    widths = [10.56, 10.56, 23.22, 12.33, 16.56, 23.22, 9.44, 12.33, 24.89]
    fill_colors = ["E1E1E1", "E1E1E1", "DBFAD7", "DBFAD7", "DBFAD7", "D1F0FF", "D1F0FF", "D1F0FF", "D1F0FF"]

    # Write the combined DataFrame back to the Excel file
    with pd.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_combined.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        # Adjust the column widths based on the content
        for i, col in enumerate(df_combined.columns):
            width = widths[i]
            worksheet.column_dimensions[chr(65 + i)].width = width

                    # Apply fill color to the entire column
            fill = PatternFill(start_color=fill_colors[i], end_color=fill_colors[i], fill_type="solid")
            for cell in worksheet[chr(65 + i)]:
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))
            
        for cell in worksheet[1]:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))